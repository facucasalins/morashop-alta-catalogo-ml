"""Alta de productos en el catálogo de MercadoLibre — MoraShop.

Lee la pestaña «Carga» del Sheet de alta, valida cada fila en estado «listo»,
toma las URLs de fotos de la pestaña «Fotos» (las sube n8n a Cloudinary),
rellena la planilla oficial de ML (descargada fresca desde Drive) y deja
el xlsx listo para subir a ML.

El script NO corrige datos: si una fila falla, escribe el motivo en
«motivo_error» y la deja en estado «error». Las filas válidas pasan a
«generado» con su número de lote.

Uso:
    python scripts/alta_catalogo_ml.py [--dry-run] [--max 24] [--categoria Suplementos]

Env:
    GOOGLE_SERVICE_ACCOUNT_JSON   (secret existente)
    ALTA_SHEET_ID                 Sheet «Paso 1 - CARGA CATALOGOS MELI»
    DRIVE_FOLDER_PLANTILLAS       carpeta donde Facu deja la planilla de ML
    DRIVE_FOLDER_SALIDA           (opcional) carpeta donde se sube el lote generado
    TELEGRAM_BOT_TOKEN / TELEGRAM_CHAT_ID  (opcionales)
    OUTPUT_DIR                    default /tmp/alta_ml
"""
from __future__ import annotations

import argparse
import io
import json
import logging
import os
import re
import sys
import unicodedata
from datetime import datetime
from zoneinfo import ZoneInfo

import gspread
import requests
from google.auth.transport.requests import AuthorizedSession
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook

log = logging.getLogger("alta_ml")
ART = ZoneInfo("America/Argentina/Buenos_Aires")
SCOPES = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]

# ---------------------------------------------------------------- helpers
def norm(s: str) -> str:
    """minúsculas, sin acentos, sin asteriscos, colapsa espacios."""
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", s.replace("*", "")).strip().lower()


def ean_valido(e: str) -> bool:
    if not re.fullmatch(r"\d{13}", e):
        return False
    suma = sum(int(d) * (1 if i % 2 == 0 else 3) for i, d in enumerate(e[:12]))
    return (10 - suma % 10) % 10 == int(e[12])


def es_num(v) -> bool:
    try:
        return float(str(v).replace(",", ".")) > 0
    except ValueError:
        return False


# Columnas de la pestaña Carga (header normalizado -> clave interna)
CARGA_COLS = {
    "categoria": "categoria", "estado": "estado", "ean": "ean", "titulo": "titulo",
    "marca": "marca", "linea": "linea", "nombre comercial": "nombre_comercial",
    "suplemento principal": "suplemento_principal", "formato": "formato",
    "formato de venta": "formato_venta", "unidades por pack": "unidades_pack",
    "peso neto": "peso_neto", "unidad de peso": "unidad_peso", "tipo de envase": "envase",
    "sabor": "sabor", "funcion": "funcion", "clase": "clase",
    "cantidad de porciones": "porciones", "descripcion": "descripcion",
    "fotos en drive": "fotos_drive", "motivo error": "motivo_error", "lote": "lote",
    "fecha generado": "fecha_generado",
}

# Encabezados de la planilla de ML (prefijo normalizado -> clave interna)
ML_COLS = {
    "titulo": "titulo", "codigo universal": "ean", "descripcion": "descripcion",
    "marca": "marca", "linea": "linea", "suplemento principal": "suplemento_principal",
    "nombre comercial": "nombre_comercial", "formato del suplemento": "formato",
    "formato de venta": "formato_venta", "peso neto": "peso_neto",
    "unidad de peso neto": "unidad_peso", "tipo de envase": "envase", "sabor": "sabor",
    "unidades por pack": "unidades_pack", "funcion del suplemento": "funcion",
    "clase de suplemento": "clase", "cantidad de porciones": "porciones",
}
ML_FILA_DATOS = 7  # la planilla de ML tiene ayudas en 1-6

# ---------------------------------------------------------------- google
def creds() -> Credentials:
    info = json.loads(os.environ["GOOGLE_SERVICE_ACCOUNT_JSON"])
    return Credentials.from_service_account_info(info, scopes=SCOPES)


def drive_buscar_plantilla(sess: AuthorizedSession, folder_id: str) -> tuple[str, str]:
    q = f"'{folder_id}' in parents and name contains 'Crear-productos' and trashed = false"
    r = sess.get("https://www.googleapis.com/drive/v3/files",
                 params={"q": q, "orderBy": "modifiedTime desc", "pageSize": 5,
                         "fields": "files(id,name,modifiedTime)",
                         "supportsAllDrives": "true", "includeItemsFromAllDrives": "true"})
    r.raise_for_status()
    files = r.json().get("files", [])
    if not files:
        raise SystemExit("❌ No hay ninguna planilla 'Crear-productos*.xlsx' en la carpeta de Drive. "
                         "Bajá una fresca desde ML (Productos de catálogo → Crear masivamente) y subila ahí.")
    f = files[0]
    log.info("Plantilla ML: %s (%s)", f["name"], f["modifiedTime"])
    return f["id"], f["name"]


def drive_descargar(sess: AuthorizedSession, file_id: str) -> bytes:
    r = sess.get(f"https://www.googleapis.com/drive/v3/files/{file_id}",
                 params={"alt": "media", "supportsAllDrives": "true"})
    r.raise_for_status()
    return r.content


def drive_subir(sess: AuthorizedSession, folder_id: str, nombre: str, data: bytes) -> str | None:
    """Best effort: si la cuota del service account lo impide, no rompe el run."""
    meta = json.dumps({"name": nombre, "parents": [folder_id]})
    files = {"metadata": ("metadata", meta, "application/json"),
             "file": (nombre, data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = sess.post("https://www.googleapis.com/upload/drive/v3/files",
                  params={"uploadType": "multipart", "supportsAllDrives": "true"}, files=files)
    if r.ok:
        return r.json().get("id")
    log.warning("No pude subir a Drive (%s): %s", r.status_code, r.text[:200])
    return None


def telegram(msg: str):
    tok, chat = os.environ.get("TELEGRAM_BOT_TOKEN"), os.environ.get("TELEGRAM_CHAT_ID")
    if not tok or not chat:
        return
    try:
        requests.post(f"https://api.telegram.org/bot{tok}/sendMessage",
                      json={"chat_id": chat, "text": msg, "parse_mode": "Markdown"}, timeout=15)
    except Exception as e:  # noqa: BLE001
        log.warning("Telegram falló: %s", e)


# ---------------------------------------------------------------- sheet
def leer_tabla(ws) -> tuple[dict[str, int], list[dict]]:
    """Devuelve (mapa clave->índice de columna 0-based, filas como dicts con _row)."""
    vals = ws.get_all_values()
    if not vals:
        return {}, []
    hdr = [norm(h) for h in vals[0]]
    filas = []
    for i, row in enumerate(vals[1:], start=2):
        row = row + [""] * (len(hdr) - len(row))
        filas.append({"_row": i, **{hdr[j]: row[j].strip() for j in range(len(hdr))}})
    return {h: j for j, h in enumerate(hdr)}, filas


def cargar_fotos(ws_fotos) -> dict[str, dict[int, str]]:
    """{ean: {imagen_n: url}} solo con ok = si."""
    _, filas = leer_tabla(ws_fotos)
    out: dict[str, dict[int, str]] = {}
    for f in filas:
        if norm(f.get("ok", "")) != "si" or not f.get("url"):
            continue
        try:
            n = int(float(f.get("imagen_n") or 1))
        except ValueError:
            n = 1
        out.setdefault(f["ean"], {})[n] = f["url"]  # si hay repetidas, gana la última (más nueva)
    return out


# ---------------------------------------------------------------- validación
def validar(p: dict, listas: dict[str, set], fotos: dict, eans_usados: set) -> list[str]:
    e = []
    ean = p["ean"].replace(" ", "").replace("-", "")
    if not ean_valido(ean):
        e.append("EAN inválido (13 dígitos con dígito verificador)")
    elif ean in eans_usados:
        e.append("EAN repetido en este lote / ya generado")
    if not p["titulo"]:
        e.append("Título vacío")
    elif len(p["titulo"]) > 200:
        e.append(f"Título de {len(p['titulo'])} caracteres (máx 200)")
    if not p["marca"]:
        e.append("Marca vacía")
    if "legend" in p["linea"].lower():
        e.append("Línea = Legend Nutrition. Línea es la del fabricante; Legend va en el Título")
    if p["linea"] and norm(p["linea"]) == norm(p["marca"]):
        e.append("Línea igual a la Marca")
    for k, nombre in (("formato", "Formato"), ("funcion", "Función"), ("clase", "Clase")):
        if not p[k]:
            e.append(f"{nombre} vacío")
        elif p[k] not in listas[k]:
            e.append(f"{nombre} «{p[k]}» no está en la lista de ML")
    if not p["sabor"]:
        e.append("Sabor vacío")
    if p["formato_venta"] and p["formato_venta"] not in listas["formato_venta"]:
        e.append(f"Formato de venta «{p['formato_venta']}» inválido (Unidad / Pack)")
    if p["formato_venta"] == "Pack" and not (es_num(p["unidades_pack"]) and float(p["unidades_pack"]) > 1):
        e.append("Pack sin «Unidades por pack» (>1)")
    if p["peso_neto"]:
        if not es_num(p["peso_neto"]):
            e.append("Peso neto no es número")
        if p["unidad_peso"] not in listas["unidad_peso"]:
            e.append("Peso neto sin unidad válida (g/kg/mg)")
    if p["porciones"] and not es_num(p["porciones"]):
        e.append("Cantidad de porciones no es número")
    if re.search(r"https?://|www\.|@|\+?\d[\d\s-]{7,}\d", p["descripcion"]):
        e.append("Descripción con link, mail o teléfono (ML lo rechaza)")
    if not fotos.get(ean, {}).get(1):
        e.append("Falta la portada en Fotos (EAN.png con ok = si). ¿Está subida y procesada?")
    return e


# ---------------------------------------------------------------- planilla ML
def mapear_columnas_ml(ws) -> dict:
    """Lee la fila 3 de la planilla ML y devuelve {clave: col_idx}, más lista de cols de imagen."""
    m, imgs = {}, []
    for c in range(1, ws.max_column + 1):
        h = norm(ws.cell(3, c).value)
        if not h:
            continue
        if h.startswith("imagen "):
            imgs.append(c)
            continue
        for pref, k in ML_COLS.items():
            if h.startswith(pref) and k not in m:
                m[k] = c
                break
    faltan = [k for k in ML_COLS.values() if k not in m]
    if faltan or len(imgs) < 1:
        raise SystemExit(f"❌ La planilla de ML no tiene las columnas esperadas. Faltan: {faltan}")
    return {**m, "_imgs": sorted(imgs)}


def listas_desde_plantilla(wb) -> dict[str, set]:
    ei = wb["extra info"]
    fila = lambda r: {str(ei.cell(r, c).value).strip() for c in range(2, ei.max_column + 1) if ei.cell(r, c).value}
    return {"formato": fila(4), "formato_venta": fila(5), "unidad_peso": fila(6),
            "funcion": fila(12), "clase": fila(13)}


def escribir_producto(ws, cols: dict, fila: int, p: dict, fotos_ean: dict[int, str]):
    def w(k, v):
        if v not in ("", None):
            ws.cell(fila, cols[k]).value = v
    w("titulo", p["titulo"]); w("ean", p["ean"]); w("descripcion", p["descripcion"])
    w("marca", p["marca"]); w("linea", p["linea"])
    w("suplemento_principal", p["suplemento_principal"]); w("nombre_comercial", p["nombre_comercial"])
    w("formato", p["formato"]); w("formato_venta", p["formato_venta"])
    w("peso_neto", float(p["peso_neto"].replace(",", ".")) if p["peso_neto"] else "")
    w("unidad_peso", p["unidad_peso"]); w("envase", p["envase"]); w("sabor", p["sabor"])
    w("unidades_pack", int(float(p["unidades_pack"])) if p["unidades_pack"] else "")
    w("funcion", p["funcion"]); w("clase", p["clase"])
    w("porciones", int(float(p["porciones"])) if p["porciones"] else "")
    for n, col in enumerate(cols["_imgs"], start=1):
        if n in fotos_ean:
            ws.cell(fila, col).value = fotos_ean[n]


# ---------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true", help="valida y genera el xlsx, pero NO toca el Sheet")
    ap.add_argument("--max", type=int, default=24, help="máximo de productos por lote (cupo de ML)")
    ap.add_argument("--categoria", default="Suplementos")
    a = ap.parse_args()
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")

    out_dir = os.environ.get("OUTPUT_DIR", "/tmp/alta_ml"); os.makedirs(out_dir, exist_ok=True)
    sheet_id = os.environ["ALTA_SHEET_ID"]
    cr = creds(); sess = AuthorizedSession(cr); gc = gspread.authorize(cr)
    sh = gc.open_by_key(sheet_id)
    ws_carga, ws_fotos = sh.worksheet("Carga"), sh.worksheet("Fotos")

    # 1) plantilla fresca de ML
    tpl_id, tpl_nombre = drive_buscar_plantilla(sess, os.environ["DRIVE_FOLDER_PLANTILLAS"])
    wb = load_workbook(io.BytesIO(drive_descargar(sess, tpl_id)))
    if a.categoria not in wb.sheetnames:
        raise SystemExit(f"❌ La planilla {tpl_nombre} no es de {a.categoria}: hojas {wb.sheetnames}")
    ws_ml = wb[a.categoria]
    cols_ml = mapear_columnas_ml(ws_ml)
    listas = listas_desde_plantilla(wb)

    # 2) datos del Sheet
    hdr_idx, filas = leer_tabla(ws_carga)
    faltan = [h for h in CARGA_COLS if h not in hdr_idx]
    if faltan:
        raise SystemExit(f"❌ A la pestaña Carga le faltan columnas: {faltan}")
    fotos = cargar_fotos(ws_fotos)
    eans_usados = {f["ean"] for f in filas if norm(f.get("estado")) in ("generado", "aprobado", "en revision")}

    validos, errores, omitidos = [], [], 0
    for f in filas:
        p = {k: f.get(h, "") for h, k in CARGA_COLS.items()}
        p["_row"] = f["_row"]
        if norm(p["estado"]) != "listo":
            continue
        if norm(p["categoria"]) != norm(a.categoria):
            omitidos += 1
            continue
        p["ean"] = p["ean"].replace(" ", "").replace("-", "")
        errs = validar(p, listas, fotos, eans_usados)
        if errs:
            errores.append((p, errs))
        else:
            validos.append(p); eans_usados.add(p["ean"])

    lote_all = validos[:a.max]
    diferidos = validos[a.max:]
    lote_id = datetime.now(ART).strftime("ML-%Y%m%d-%H%M")
    ahora = datetime.now(ART).isoformat(timespec="seconds")

    # 3) rellenar planilla
    for i, p in enumerate(lote_all):
        escribir_producto(ws_ml, cols_ml, ML_FILA_DATOS + i, p, fotos.get(p["ean"], {}))
    salida = os.path.join(out_dir, f"{lote_id}_{a.categoria}_{len(lote_all)}prod.xlsx")
    wb.save(salida)

    # 4) reporte
    rep = [f"# Lote {lote_id} — {a.categoria}", f"Plantilla: {tpl_nombre}", "",
           f"✅ Incluidos: {len(lote_all)}   ❌ Con error: {len(errores)}   "
           f"⏸ Diferidos por cupo: {len(diferidos)}   ↷ Otra categoría: {omitidos}", ""]
    rep += [f"- fila {p['_row']}  {p['ean']}  {p['titulo'][:60]}" for p in lote_all]
    if errores:
        rep += ["", "## Errores (quedan en estado «error»)"]
        rep += [f"- fila {p['_row']}  {p['ean'] or '(sin EAN)'}: " + " | ".join(e) for p, e in errores]
    if diferidos:
        rep += ["", f"## Diferidos (siguen en «listo», entran en el próximo lote)"]
        rep += [f"- fila {p['_row']}  {p['ean']}" for p in diferidos]
    rep_txt = "\n".join(rep)
    open(os.path.join(out_dir, f"{lote_id}_reporte.md"), "w").write(rep_txt)
    print(rep_txt)

    # 5) escribir estados en el Sheet
    if a.dry_run:
        log.info("DRY RUN: no toco el Sheet")
    else:
        col = lambda k: hdr_idx[next(h for h, kk in CARGA_COLS.items() if kk == k)] + 1
        upd = []
        for p in lote_all:
            upd += [{"range": gspread.utils.rowcol_to_a1(p["_row"], col("estado")), "values": [["generado"]]},
                    {"range": gspread.utils.rowcol_to_a1(p["_row"], col("motivo_error")), "values": [[""]]},
                    {"range": gspread.utils.rowcol_to_a1(p["_row"], col("lote")), "values": [[lote_id]]},
                    {"range": gspread.utils.rowcol_to_a1(p["_row"], col("fecha_generado")), "values": [[ahora]]}]
        for p, e in errores:
            upd += [{"range": gspread.utils.rowcol_to_a1(p["_row"], col("estado")), "values": [["error"]]},
                    {"range": gspread.utils.rowcol_to_a1(p["_row"], col("motivo_error")), "values": [[" | ".join(e)]]}]
        if upd:
            ws_carga.batch_update(upd)
            log.info("Sheet actualizado: %d filas", len(lote_all) + len(errores))
        if lote_all and os.environ.get("DRIVE_FOLDER_SALIDA"):
            fid = drive_subir(sess, os.environ["DRIVE_FOLDER_SALIDA"], os.path.basename(salida), open(salida, "rb").read())
            if fid:
                log.info("Subido a Drive: %s", fid)

    telegram(f"📦 *Alta ML {lote_id}*\n✅ {len(lote_all)} listos para subir · ❌ {len(errores)} con error · "
             f"⏸ {len(diferidos)} diferidos{' · (dry run)' if a.dry_run else ''}\n"
             f"Bajá el xlsx del run de GitHub Actions.")
    if not lote_all:
        log.warning("No hay productos válidos para este lote.")
        sys.exit(2)


if __name__ == "__main__":
    main()
